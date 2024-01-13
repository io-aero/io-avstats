# Copyright (c) 2022-2023 IO-Aero. All rights reserved. Use of this
# source code is governed by the IO-Aero License, that can
# be found in the LICENSE.md file.

"""Managing the database schema of the PostgreSQL database."""
import json
import logging
import os
import zipfile
from datetime import datetime
from datetime import timezone

import requests
from iocommon import db_utils
from iocommon import io_config
from iocommon import io_glob
from iocommon import io_utils
from openpyxl.reader.excel import load_workbook
from psycopg2.errors import ForeignKeyViolation  # pylint: disable=no-name-in-module
from psycopg2.errors import UniqueViolation  # pylint: disable=no-name-in-module
from psycopg2.extensions import connection
from psycopg2.extensions import cursor

from ioavstats import glob
from ioavstats import utils

# -----------------------------------------------------------------------------
# Global variables.
# -----------------------------------------------------------------------------

IO_LAST_SEEN = datetime.now(timezone.utc)

logger = logging.getLogger(__name__)


# ------------------------------------------------------------------
# Load airport data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-lines
# pylint: disable=too-many-statements
def _load_airport_data() -> None:
    logger.debug(io_glob.LOGGER_START)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_airports
    """,
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()
        io_utils.progress_msg("-" * 80)
        # INFO.00.087 Database table io_airports: Delete the existing data
        io_utils.progress_msg(glob.INFO_00_087)
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")

    # ------------------------------------------------------------------
    # Start processing airport data.
    # ------------------------------------------------------------------

    us_states: list[str] = _sql_query_us_states(conn_pg)

    locids: list[str] = _load_npias_data(us_states)

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_airports_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.943 The airport file '{filename}' is missing
        io_utils.terminate_fatal(
            glob.ERROR_00_943.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0
    count_upsert = 0
    count_usable = 0

    x_idx = 0
    y_idx = 1
    global_idx = 3
    ident_idx = 4
    name_idx = 5
    latitude_idx = 6
    longitude_idx = 7
    elevation_idx = 8
    type_code_idx = 10
    servcity_idx = 11
    state_idx = 12
    country_idx = 13
    operstatus_idx = 14
    privateuse_idx = 15
    iapexists_idx = 16
    dodhiflib_idx = 17
    far91_idx = 18
    far93_idx = 19
    mil_code_idx = 20
    airanal_idx = 21

    # ------------------------------------------------------------------
    # Load the airport data.
    # ------------------------------------------------------------------

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        count_select += 1

        ident = row[ident_idx].value
        if ident is None or ident not in locids:
            continue

        country = row[country_idx].value
        if country != "UNITED STATES":
            continue

        dec_longitude = row[x_idx].value
        if dec_longitude == "X":
            continue

        state = row[state_idx].value
        if state is None or state not in us_states:
            continue

        count_usable += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        airanal = row[airanal_idx].value
        country = "USA"
        dec_latitude = row[y_idx].value
        dodhiflib = row[dodhiflib_idx].value
        elevation = row[elevation_idx].value
        far91 = row[far91_idx].value
        far93 = row[far93_idx].value
        global_id = row[global_idx].value
        iapexists = row[iapexists_idx].value
        latitude = row[latitude_idx].value
        longitude = row[longitude_idx].value
        mil_code = row[mil_code_idx].value
        name = row[name_idx].value
        operstatus = row[operstatus_idx].value
        privateuse = row[privateuse_idx].value
        servcity = row[servcity_idx].value
        state = row[state_idx].value
        type_code = row[type_code_idx].value

        cur_pg.execute(
            """
        INSERT INTO io_airports AS ia (
               global_id,
               airanal,
               country,
               dec_latitude,
               dec_longitude,
               dodhiflib,
               elevation,
               far91,
               far93,
               iapexists,
               ident,
               latitude,
               longitude,
               mil_code,
               name,
               operstatus,
               privateuse,
               servcity,
               state,
               type_code,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s,%s,
               %s
               )
        ON CONFLICT ON CONSTRAINT io_airports_pkey
        DO UPDATE
           SET airanal = %s,
               country = %s,
               dec_latitude = %s,
               dec_longitude = %s,
               dodhiflib = %s,
               elevation = %s,
               far91 = %s,
               far93 = %s,
               iapexists = %s,
               ident = %s,
               latitude = %s,
               longitude = %s,
               mil_code = %s,
               name = %s,
               operstatus = %s,
               privateuse = %s,
               servcity = %s,
               state = %s,
               type_code = %s,
               last_processed = %s
         WHERE ia.global_id = %s
        """,
            (
                global_id,
                airanal,
                country,
                dec_latitude,
                dec_longitude,
                dodhiflib,
                elevation,
                far91,
                far93,
                iapexists,
                ident,
                latitude,
                longitude,
                mil_code,
                name,
                operstatus,
                privateuse,
                servcity,
                state,
                type_code,
                datetime.now(),
                airanal,
                country,
                dec_latitude,
                dec_longitude,
                dodhiflib,
                elevation,
                far91,
                far93,
                iapexists,
                ident,
                latitude,
                longitude,
                mil_code,
                name,
                operstatus,
                privateuse,
                servcity,
                state,
                type_code,
                datetime.now(),
                global_id,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load aviation occurrence categories from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def _load_aviation_occurrence_categories() -> None:
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_aviation_occurrence_categories_xlsx.replace(
            "/", os.sep
        ),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.937 The aviation occurrence categories file '{filename}' is missing
        io_utils.terminate_fatal(
            glob.ERROR_00_937.replace("{filename}", filename_excel)
        )

    # INFO.00.074 Database table io_aviation_occurrence_categories: Load data from file '{filename}'
    io_utils.progress_msg(
        glob.INFO_00_074.replace(
            "{filename}",
            io_config.settings.download_file_aviation_occurrence_categories_xlsx,
        )
    )
    io_utils.progress_msg("-" * 80)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    count_delete = 0
    count_upsert = 0
    count_select = 0

    # ------------------------------------------------------------------
    # Load the aviation occurrence categories.
    # ------------------------------------------------------------------

    cictt_code_idx = 0
    identifier_idx = 1
    definition_idx = 2

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        cictt_code = row[cictt_code_idx].value.upper().rstrip()
        if cictt_code == "CICTT CODE":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        identifier = row[identifier_idx].value.upper().rstrip()
        definition = row[definition_idx].value.rstrip()

        cur_pg.execute(
            """
        INSERT INTO io_aviation_occurrence_categories AS aoc (
               cictt_code,
               identifier,
               definition,
               first_processed,
               last_seen
               ) VALUES (
               %s,%s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_aviation_occurrence_categories_pkey
        DO UPDATE
           SET identifier = %s,
               definition = %s,
               last_processed = %s,
               last_seen = %s
         WHERE aoc.cictt_code = %s
        """,
            (
                cictt_code,
                identifier,
                definition,
                datetime.now(),
                IO_LAST_SEEN,
                identifier,
                definition,
                datetime.now(),
                IO_LAST_SEEN,
                cictt_code,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Delete the obsolete data.
    # ------------------------------------------------------------------

    count_select = 0

    # pylint: disable=line-too-long
    cur_pg.execute(
        """
    SELECT cictt_code
      FROM io_aviation_occurrence_categories
     WHERE last_seen <> %s
        """,
        (IO_LAST_SEEN,),
    )

    for row_pg in cur_pg.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cictt_code = row_pg["cictt_code"]  # type: ignore

        try:
            cur_pg.execute(
                """
            DELETE FROM io_aviation_occurrence_categories
             WHERE cictt_code = %s;
            """,
                (cictt_code,),
            )
            if cur_pg.rowcount > 0:
                count_delete += cur_pg.rowcount
                io_utils.progress_msg(f"Deleted cictt_code={cictt_code}")
        except ForeignKeyViolation:
            io_utils.progress_msg(f"Failed to delete cictt_code={cictt_code}")

    conn_pg.commit()

    if count_select > 0:
        io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    if count_delete > 0:
        io_utils.progress_msg(f"Number rows deleted  : {str(count_delete):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    utils.upd_io_processed_files(
        io_config.settings.download_file_aviation_occurrence_categories_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load country data from a JSON file.
# ------------------------------------------------------------------
def _load_country_data(
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    logger.debug(io_glob.LOGGER_START)

    filename_json = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_countries_states_json.replace("/", os.sep),
    )

    if not os.path.isfile(filename_json):
        # ERROR.00.934 The country and state data file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_934.replace("{filename}", filename_json))

    count_insert = 0
    count_select = 0
    count_update = 0

    with open(filename_json, "r", encoding=io_glob.FILE_ENCODING_DEFAULT) as input_file:
        input_data = json.load(input_file)

        for record in input_data:
            count_select += 1

            if record["type"] == "country":
                try:
                    cur_pg.execute(
                        """
                    INSERT INTO io_countries (
                           country,country_name,dec_latitude,dec_longitude,first_processed
                           ) VALUES (
                           %s,%s,%s,%s,%s);
                    """,
                        (
                            record["country"],
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                        ),
                    )
                    count_insert += 1
                except UniqueViolation:
                    # pylint: disable=line-too-long
                    cur_pg.execute(
                        """
                    UPDATE io_countries SET
                           country_name = %s,
                           dec_latitude = %s,
                           dec_longitude = %s,
                           last_processed = %s
                     WHERE country = %s
                       AND NOT (country_name = %s
                            AND dec_latitude = %s
                            AND dec_longitude = %s);
                    """,
                        (
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                            record["country"],
                            record["country_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                        ),
                    )
                    count_update += cur_pg.rowcount

    utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load NPIAS data from an MS Excel file.
# ------------------------------------------------------------------
def _load_npias_data(us_states) -> list[str]:
    # ------------------------------------------------------------------
    # Start processing NPIAS data.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_npias_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.945 The NPIAS file '{filename}' is missing
        io_utils.terminate_fatal(
            glob.ERROR_00_945.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0
    count_usable = 0

    state_idx = 0
    locid_idx = 3

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    locids: list[str] = []

    # pylint: disable=R0801
    for row in workbook.active:
        count_select += 1
        if not row[state_idx].value in us_states:
            continue

        locids.append(row[locid_idx].value)

        count_usable += 1

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg(f"Number rows usable   : {str(count_usable):>8}")

    return locids


# ------------------------------------------------------------------
# Load runway data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def _load_runway_data() -> None:
    logger.debug(io_glob.LOGGER_START)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Load airport identifications.
    # ------------------------------------------------------------------
    # INFO.00.088 Database table io_airports: Load the global identifications
    io_utils.progress_msg("-" * 80)
    io_utils.progress_msg(glob.INFO_00_088)

    count_select = 0

    cur_pg.execute(
        """
    SELECT global_id
      FROM io_airports
     ORDER BY 1;
    """,
    )

    runway_data: dict[str, tuple[str | None, float | None]] = {}

    for row in cur_pg:
        count_select += 1
        runway_data[row[0]] = (None, None)

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Start processing airport data.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_faa_runways_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.944 The runway file '{filename}' is missing
        io_utils.terminate_fatal(
            glob.ERROR_00_944.replace("{filename}", filename_excel)
        )

    # INFO.00.089 Database table io_airports: Load data from file '{filename}'
    io_utils.progress_msg(glob.INFO_00_089.replace("{filename}", filename_excel))

    count_select = 0

    airport_id_idx = 2
    comp_code_idx = 7
    dim_uom_idx = 6
    length_idx = 4

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        airport_id = row[airport_id_idx].value
        if airport_id not in runway_data:
            continue

        count_select += 1

        (
            comp_code,
            length,
        ) = runway_data[airport_id]

        new_length = row[length_idx].value

        dim_vom = row[dim_uom_idx].value
        if dim_vom == "M":
            new_length = new_length * 3.28084

        if length is None or new_length > length:
            runway_data[airport_id] = (
                row[comp_code_idx].value,
                new_length,
            )

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load the runway data.
    # ------------------------------------------------------------------

    # INFO.00.090 Database table io_airports: Update the runway data
    io_utils.progress_msg(glob.INFO_00_090)

    count_select = 0
    count_update = 0

    # pylint: disable=R0801
    for airport_id, (comp_code, length) in runway_data.items():
        count_select += 1

        if comp_code is None:
            continue

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cur_pg.execute(
            """
        UPDATE io_airports ia
           SET max_runway_comp_code = %s,
               max_runway_length = %s,
               last_processed = %s
         WHERE ia.global_id = %s
        """,
            (
                comp_code,
                length,
                datetime.now(),
                airport_id,
            ),
        )
        count_update += cur_pg.rowcount

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load sequence of events sequence data from an MS Excel file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-statements
def _load_sequence_of_events() -> None:
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    filename_excel = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_sequence_of_events_xlsx.replace("/", os.sep),
    )

    if not os.path.isfile(filename_excel):
        # ERROR.00.938 The sequence of events file '{filename}' is missing
        io_utils.terminate_fatal(
            glob.ERROR_00_938.replace("{filename}", filename_excel)
        )

    # INFO.00.076 Database table io_sequence_of_events: Load data from file '{filename}'
    io_utils.progress_msg(
        glob.INFO_00_076.replace(
            "{filename}", io_config.settings.download_file_sequence_of_events_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    count_delete = 0
    count_upsert = 0
    count_select = 0

    eventsoe_no_idx = 0
    meaning_idx = 1
    cictt_code_idx = 2

    # ------------------------------------------------------------------
    # Load the sequence of events data.
    # ------------------------------------------------------------------

    cictt_codes = _sql_query_cictt_codes(conn_pg)

    workbook = load_workbook(
        filename=filename_excel,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        soe_no = str(row[eventsoe_no_idx].value).rjust(3, "0")
        if soe_no == "eventsoe_no":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        cictt_code = (
            row[cictt_code_idx].value.upper().rstrip()
            if row[cictt_code_idx].value
            else None
        )
        if cictt_code is None or cictt_code not in cictt_codes:
            continue

        meaning = row[meaning_idx].value.rstrip()

        cur_pg.execute(
            """
        INSERT INTO io_sequence_of_events AS isoe (
               soe_no,
               meaning,
               cictt_code,
               first_processed,
               last_seen
               ) VALUES (
               %s,%s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_sequence_of_events_pkey
        DO UPDATE
           SET meaning = %s,
               cictt_code = %s,
               last_processed = %s,
               last_seen = %s
         WHERE isoe.soe_no = %s
        """,
            (
                soe_no,
                meaning,
                cictt_code if cictt_code else None,
                datetime.now(),
                IO_LAST_SEEN,
                meaning,
                cictt_code,
                datetime.now(),
                IO_LAST_SEEN,
                soe_no,
            ),
        )
        count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    # ------------------------------------------------------------------
    # Delete the obsolete data.
    # ------------------------------------------------------------------

    count_select = 0

    # pylint: disable=line-too-long
    cur_pg.execute(
        """
    SELECT soe_no
      FROM io_sequence_of_events
     WHERE last_seen <> %s
        """,
        (IO_LAST_SEEN,),
    )

    for row_pg in cur_pg.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        soe_no = row_pg["soe_no"]  # type: ignore

        try:
            cur_pg.execute(
                """
            DELETE FROM io_sequence_of_events
             WHERE soe_no = %s;
            """,
                (soe_no,),
            )
            if cur_pg.rowcount > 0:
                count_delete += cur_pg.rowcount
                io_utils.progress_msg(f"Deleted soe_no={soe_no}")
        except ForeignKeyViolation:
            io_utils.progress_msg(f"Failed to delete soe_no={soe_no}")

    conn_pg.commit()

    if count_select > 0:
        io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")
    if count_delete > 0:
        io_utils.progress_msg(f"Number rows deleted  : {str(count_delete):>8}")

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    utils.upd_io_processed_files(
        io_config.settings.download_file_sequence_of_events_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load city data from a US city file.
# ------------------------------------------------------------------
# pylint: disable=too-many-locals
def _load_simplemaps_data_cities_from_us_cities(
    conn_pg: connection, cur_pg: cursor
) -> None:
    logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_cities_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.914 The US city file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_914.replace("{filename}", filename))

    # INFO.00.027 Database table io_lat_lng: Load city data from file '{filename}'
    io_utils.progress_msg(
        glob.INFO_00_027.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_cities_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_upsert = 0
    count_select = 0

    city_idx = 0
    lat_idx = 6
    lng_idx = 7
    state_idx = 2

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        city = row[city_idx].value.upper().rstrip()
        if city == "CITY":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()

        # pylint: disable=line-too-long
        cur_pg.execute(
            """
        INSERT INTO io_lat_lng AS ill (
               type,
               country,
               state,
               city,
               dec_latitude,
               dec_longitude,
               source,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
        DO UPDATE
           SET dec_latitude = %s,
               dec_longitude = %s,
               source = %s,
               last_processed = %s
         WHERE ill.type = %s
           AND ill.country = %s
           AND ill.state = %s
           AND ill.city = %s
           AND NOT (
               ill.dec_latitude = %s
           AND ill.dec_longitude = %s
           );
        """,
            (
                glob.IO_LAT_LNG_TYPE_CITY,
                glob.COUNTRY_USA,
                state,
                city,
                lat,
                lng,
                glob.SOURCE_SM_US_CITIES,
                datetime.now(),
                lat,
                lng,
                glob.SOURCE_SM_US_CITIES,
                datetime.now(),
                glob.IO_LAT_LNG_TYPE_CITY,
                glob.COUNTRY_USA,
                state,
                city,
                lat,
                lng,
            ),
        )
        count_upsert += cur_pg.rowcount

    utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_cities_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load zip data from a US city file.
# ------------------------------------------------------------------
# pylint: disable=R0801
# pylint: disable=too-many-locals
def _load_simplemaps_data_zips_from_us_cities(
    conn_pg: connection, cur_pg: cursor
) -> None:
    logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_cities_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.914 The US city file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_914.replace("{filename}", filename))

    # INFO.00.039 Database table io_lat_lng: Load zipcode data from file '{filename}'
    io_utils.progress_msg(
        glob.INFO_00_039.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_cities_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_insert = 0
    count_select = 0

    city_idx = 0
    lat_idx = 6
    lng_idx = 7
    state_idx = 2
    zipcodes_idx = 15

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    for row in workbook.active:
        city = row[city_idx].value.upper().rstrip()
        if city == "CITY":
            continue

        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()
        zipcodes = list(f"{row[zipcodes_idx].value}".split(" "))

        for zipcode in zipcodes:
            count_select += 1
            if count_select % io_config.settings.database_commit_size == 0:
                conn_pg.commit()
                io_utils.progress_msg(
                    f"Number of rows so far read : {str(count_select):>8}"
                )

            # pylint: disable=line-too-long
            cur_pg.execute(
                """
            INSERT INTO io_lat_lng (
                   type,
                   country,
                   state,
                   city,
                   zipcode,
                   dec_latitude,
                   dec_longitude,
                   source,
                   first_processed
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s,%s,%s,%s)
            ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
            DO NOTHING;
            """,
                (
                    glob.IO_LAT_LNG_TYPE_ZIPCODE,
                    glob.COUNTRY_USA,
                    state,
                    city,
                    zipcode.rstrip(),
                    lat,
                    lng,
                    glob.SOURCE_SM_US_CITIES,
                    datetime.now(),
                ),
            )
            count_insert += cur_pg.rowcount

    utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_cities_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load zip data from a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-locals
def _load_simplemaps_data_zips_from_us_zips(
    conn_pg: connection, cur_pg: cursor
) -> None:
    logger.debug(io_glob.LOGGER_START)

    filename = io_config.settings.download_file_simplemaps_us_zips_xlsx

    if not os.path.isfile(filename):
        # ERROR.00.913 The US zip code file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_913.replace("{filename}", filename))

    # INFO.00.025 Database table io_lat_lng: Load zipcode data rom file '{filename}'
    io_utils.progress_msg(
        glob.INFO_00_025.replace(
            "{filename}", io_config.settings.download_file_simplemaps_us_zips_xlsx
        )
    )
    io_utils.progress_msg("-" * 80)

    count_duplicates = 0
    count_upsert = 0
    count_select = 0

    zip_idx = 0
    lat_idx = 1
    lng_idx = 2
    city_idx = 3
    state_idx = 4

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    for row in workbook.active:
        zipcode = f"{row[zip_idx].value:05}".rstrip()
        if zipcode == "zip00":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        city = row[city_idx].value.upper().rstrip()
        lat = row[lat_idx].value
        lng = row[lng_idx].value
        state = row[state_idx].value.upper().rstrip()

        # pylint: disable=line-too-long
        cur_pg.execute(
            """
        INSERT INTO io_lat_lng AS ill (
               type,
               country,
               state,
               city,
               zipcode,
               dec_latitude,
               dec_longitude,
               source,
               first_processed
               ) VALUES (
               %s,%s,%s,%s,%s,
               %s,%s,%s,%s
               )
        ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
        DO UPDATE
           SET dec_latitude = %s,
               dec_longitude = %s,
               source = %s,
               last_processed = %s
         WHERE ill.type = %s
           AND ill.country = %s
           AND ill.state = %s
           AND ill.city = %s
           AND ill.zipcode = %s
           AND NOT (
               ill.dec_latitude = %s
           AND ill.dec_longitude = %s
           );
        """,
            (
                glob.IO_LAT_LNG_TYPE_ZIPCODE,
                glob.COUNTRY_USA,
                state,
                city,
                zipcode,
                lat,
                lng,
                glob.SOURCE_SM_US_ZIP_CODES,
                datetime.now(),
                lat,
                lng,
                glob.SOURCE_SM_US_ZIP_CODES,
                datetime.now(),
                glob.IO_LAT_LNG_TYPE_ZIPCODE,
                glob.COUNTRY_USA,
                state,
                city,
                zipcode,
                lat,
                lng,
            ),
        )
        count_upsert += cur_pg.rowcount

    utils.upd_io_processed_files(
        io_config.settings.download_file_simplemaps_us_zips_xlsx, cur_pg
    )

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    if count_duplicates > 0:
        io_utils.progress_msg(f"Number rows duplicate: {str(count_duplicates):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load state data from a JSON file.
# ------------------------------------------------------------------
def _load_state_data(
    conn_pg: connection,
    cur_pg: cursor,
) -> None:
    logger.debug(io_glob.LOGGER_START)

    filename_json = os.path.join(
        os.getcwd(),
        io_config.settings.download_file_countries_states_json.replace("/", os.sep),
    )

    if not os.path.isfile(filename_json):
        # ERROR.00.934 The country and state data file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_934.replace("{filename}", filename_json))

    count_insert = 0
    count_select = 0
    count_update = 0

    with open(filename_json, "r", encoding=io_glob.FILE_ENCODING_DEFAULT) as input_file:
        input_data = json.load(input_file)

        for record in input_data:
            count_select += 1

            if record["type"] == "state":
                try:
                    cur_pg.execute(
                        """
                    INSERT INTO io_states (
                           country,state, state_name,dec_latitude,dec_longitude,first_processed
                           ) VALUES (%s,%s,%s,%s,%s,%s);
                    """,
                        (
                            record["country"],
                            record["state"],
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                        ),
                    )
                    count_insert += 1
                except UniqueViolation:
                    cur_pg.execute(
                        """
                    UPDATE io_states SET
                           state_name = %s,dec_latitude = %s,dec_longitude = %s,last_processed = %s
                     WHERE
                           country = %s AND state = %s
                       AND NOT (state_name = %s AND dec_latitude = %s AND dec_longitude = %s);
                    """,
                        (
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                            datetime.now(),
                            record["country"],
                            record["state"],
                            record["state_name"],
                            record["dec_latitude"],
                            record["dec_longitude"],
                        ),
                    )
                    count_update += cur_pg.rowcount

    utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_update > 0:
        io_utils.progress_msg(f"Number rows updated  : {str(count_update):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Determine and load city averages.
# ------------------------------------------------------------------
def _load_table_io_lat_lng_average(conn_pg, cur_pg) -> None:
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Delete averaged data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
     WHERE
           source = %s;
    """,
        (glob.SOURCE_AVERAGE,),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            glob.INFO_00_063.replace("{data_source}", glob.SOURCE_AVERAGE)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Insert averaged data.
    # ------------------------------------------------------------------
    # INFO.00.062 Database table io_lat_lng: Load the averaged city data
    io_utils.progress_msg(glob.INFO_00_062)
    io_utils.progress_msg("-" * 80)

    count_duplicates = 0
    count_insert = 0
    count_select = 0

    conn_pg_2, cur_pg_2 = db_utils.get_postgres_cursor()

    conn_pg_2.set_session(autocommit=False)

    # pylint: disable=line-too-long
    cur_pg_2.execute(
        f"""
    SELECT country,
           state,
           city,
           sum(dec_latitude)/count(*) dec_latitude,
           sum(dec_longitude)/count(*) dec_longitude
      FROM io_lat_lng
     WHERE type = '{glob.IO_LAT_LNG_TYPE_ZIPCODE}'
     GROUP BY country, state, city;
        """,
    )

    for row_pg in cur_pg_2.fetchall():
        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        try:
            cur_pg.execute(
                """
            INSERT INTO io_lat_lng (
                   type,
                   country,
                   state,
                   city,
                   dec_latitude,
                   dec_longitude,
                   source,
                   first_processed
                   ) VALUES (
                   %s,%s,%s,%s,%s,
                   %s,%s,%s
                   )
            ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
            DO NOTHING;
            """,
                (
                    glob.IO_LAT_LNG_TYPE_CITY,
                    row_pg["country"],  # type: ignore
                    row_pg["state"],  # type: ignore
                    row_pg["city"],  # type: ignore
                    row_pg["dec_latitude"],  # type: ignore
                    row_pg["dec_longitude"],  # type: ignore
                    glob.SOURCE_AVERAGE,
                    datetime.now(),
                ),
            )
            count_insert += 1
        except UniqueViolation:
            count_duplicates += cur_pg.rowcount

    conn_pg.commit()

    cur_pg_2.close()
    conn_pg_2.close()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_insert > 0:
        io_utils.progress_msg(f"Number rows inserted : {str(count_insert):>8}")

    if count_duplicates > 0:
        io_utils.progress_msg(f"Number rows duplicate: {str(count_duplicates):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load ZIP Code Database data.
# ------------------------------------------------------------------
def _load_zip_codes_org_data() -> None:
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------
    filename_xlsx = io_config.settings.download_file_zip_codes_org_xlsx

    if not os.path.isfile(filename_xlsx):
        # ERROR.00.935 The Zip Code Database file '{filename}' is missing
        io_utils.terminate_fatal(glob.ERROR_00_935.replace("{filename}", filename_xlsx))

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    conn_pg.set_session(autocommit=False)

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
     WHERE SOURCE = %s;
    """,
        (glob.SOURCE_ZCO_ZIP_CODES,),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            glob.INFO_00_063.replace("{data_source}", glob.SOURCE_ZCO_ZIP_CODES)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Insert new data.
    # ------------------------------------------------------------------
    _load_zip_codes_org_data_zips(conn_pg, cur_pg, filename_xlsx)
    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Delete and insert averaged data.
    # ------------------------------------------------------------------
    _load_table_io_lat_lng_average(conn_pg, cur_pg)

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------
    utils.upd_io_processed_files(
        io_config.settings.download_file_zip_codes_org_xlsx, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load the estimated zip code data from the ZIP Code Database
# into the PostgreSQL database.
# ------------------------------------------------------------------
def _load_zip_codes_org_data_zips(conn_pg, cur_pg, filename) -> None:
    logger.debug(io_glob.LOGGER_START)

    # INFO.00.061 Database table io_lat_lng: Load the estimated zip code data
    io_utils.progress_msg(glob.INFO_00_061)
    io_utils.progress_msg("-" * 80)

    count_upsert = 0
    count_select = 0

    zip_idx = 0
    type_idx = 1
    primary_city_idx = 3
    acceptable_cities_idx = 4
    state_idx = 6
    country_idx = 11
    latitude_idx = 12
    longitude_idx = 13

    workbook = load_workbook(
        filename=filename,
        read_only=True,
        data_only=True,
    )

    # pylint: disable=R0801
    for row in workbook.active:
        zipcode = f"{row[zip_idx].value:05}".rstrip()
        if zipcode == "zip":
            continue

        count_select += 1

        if count_select % io_config.settings.database_commit_size == 0:
            conn_pg.commit()
            io_utils.progress_msg(
                f"Number of rows so far read : {str(count_select):>8}"
            )

        if row[type_idx].value != "STANDARD" or row[country_idx].value != "US":
            continue

        primary_city = [row[primary_city_idx].value.upper().rstrip()]
        acceptable_cities = (
            row[acceptable_cities_idx].value.upper().split(",")
            if row[acceptable_cities_idx].value
            else []
        )
        state = row[state_idx].value.upper().rstrip()
        lat = row[latitude_idx].value
        lng = row[longitude_idx].value

        for city in acceptable_cities:
            primary_city.append(city.rstrip())

        for city in primary_city:
            if city and city.rstrip():
                # pylint: disable=line-too-long
                cur_pg.execute(
                    """
                INSERT INTO io_lat_lng AS ill (
                       type,
                       country,
                       state,
                       city,
                       zipcode,
                       dec_latitude,
                       dec_longitude,
                       source,
                       first_processed
                       ) VALUES (
                       %s,%s,%s,%s,%s,
                       %s,%s,%s,%s
                       )
                ON CONFLICT ON CONSTRAINT io_lat_lng_type_country_state_city_zipcode_key
                DO UPDATE
                   SET dec_latitude = %s,
                       dec_longitude = %s,
                       source = %s,
                       last_processed = %s
                 WHERE ill.type = %s
                   AND ill.country = %s
                   AND ill.state = %s
                   AND ill.city = %s
                   AND ill.zipcode = %s
                   AND ill.source = %s
                   AND NOT (
                       ill.dec_latitude = %s
                   AND ill.dec_longitude = %s
                   );
                """,
                    (
                        glob.IO_LAT_LNG_TYPE_ZIPCODE,
                        glob.COUNTRY_USA,
                        state.rstrip(),
                        city.rstrip(),
                        zipcode.rstrip(),
                        lat,
                        lng,
                        glob.SOURCE_ZCO_ZIP_CODES,
                        datetime.now(),
                        lat,
                        lng,
                        glob.SOURCE_ZCO_ZIP_CODES,
                        datetime.now(),
                        glob.IO_LAT_LNG_TYPE_ZIPCODE,
                        glob.COUNTRY_USA,
                        state.rstrip(),
                        city.rstrip(),
                        zipcode.rstrip(),
                        glob.SOURCE_SM_US_CITIES,
                        lat,
                        lng,
                    ),
                )
                count_upsert += cur_pg.rowcount

    workbook.close()

    conn_pg.commit()

    io_utils.progress_msg(f"Number rows selected : {str(count_select):>8}")

    if count_upsert > 0:
        io_utils.progress_msg(f"Number rows upserted : {str(count_upsert):>8}")

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Execute a query that returns the list of CICTT codes.
# ------------------------------------------------------------------
def _sql_query_cictt_codes(conn_pg: connection) -> list[str]:
    with conn_pg.cursor() as cur:  # type: ignore
        cur.execute(
            """
        SELECT cictt_code
          FROM io_aviation_occurrence_categories;
        """
        )

        data = []

        for row in cur:
            data.append(row[0])

        return sorted(data)


# ------------------------------------------------------------------
# Execute a query that returns the list of US states.
# ------------------------------------------------------------------
def _sql_query_us_states(conn_pg: connection) -> list[str]:
    with conn_pg.cursor() as cur:  # type: ignore
        cur.execute(
            """
        SELECT state
          FROM io_states
         WHERE country = 'USA';
        """
        )

        data = []

        for row in cur:
            data.append(row[0])

        return sorted(data)


# ------------------------------------------------------------------
# Download a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_us_cities_file() -> None:
    """Download a US zip code file."""
    logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_simplemaps_us_cities

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                glob.ERROR_00_906.replace("{status_code}", str(file_resp.status_code))
            )

        # INFO.00.030 The connection to the US city file '{filename}'
        # on the simplemaps download page was successfully established
        io_utils.progress_msg(
            glob.INFO_00_030.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_zip = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_simplemaps_us_cities_zip,
        )

        no_chunks = 0

        with open(filename_zip, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            ).replace("{no_chunks}", str(no_chunks))
        )

        try:
            zipped_files = zipfile.ZipFile(  # pylint: disable=consider-using-with
                filename_zip
            )

            for zipped_file in zipped_files.namelist():
                zipped_files.extract(zipped_file, io_config.settings.download_work_dir)

            zipped_files.close()
        except zipfile.BadZipFile:
            # ERROR.00.907 File '{filename}' is not a zip file
            io_utils.terminate_fatal(
                glob.ERROR_00_907.replace("{filename}", filename_zip)
            )

        os.remove(filename_zip)
        # INFO.00.024 The file '{filename}'  was successfully unpacked
        io_utils.progress_msg(
            glob.INFO_00_024.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_cities_zip
            )
        )
    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Download a US zip code file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_us_zips_file() -> None:
    """Download a US zip code file."""
    logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_simplemaps_us_zips

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                glob.ERROR_00_906.replace("{status_code}", str(file_resp.status_code))
            )

        # INFO.00.022 The connection to the US zip code file '{filename}'
        # on the simplemaps download page was successfully established
        io_utils.progress_msg(
            glob.INFO_00_022.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_zip = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_simplemaps_us_zips_zip,
        )

        no_chunks = 0

        with open(filename_zip, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            ).replace("{no_chunks}", str(no_chunks))
        )

        try:
            zipped_files = zipfile.ZipFile(  # pylint: disable=consider-using-with
                filename_zip
            )

            for zipped_file in zipped_files.namelist():
                zipped_files.extract(zipped_file, io_config.settings.download_work_dir)

            zipped_files.close()
        except zipfile.BadZipFile:
            # ERROR.00.907 File '{filename}' is not a zip file
            io_utils.terminate_fatal(
                glob.ERROR_00_907.replace("{filename}", filename_zip)
            )

        os.remove(filename_zip)
        # INFO.00.024 The file '{filename}'  was successfully unpacked
        io_utils.progress_msg(
            glob.INFO_00_024.replace(
                "{filename}", io_config.settings.download_file_simplemaps_us_zips_zip
            )
        )
    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Download the ZIP Code Database file.
# ------------------------------------------------------------------
# pylint: disable=too-many-branches
# pylint: disable=too-many-locals
# pylint: disable=too-many-statements
def download_zip_code_db_file() -> None:
    """Download the ZIP Code Database file."""
    logger.debug(io_glob.LOGGER_START)

    url = io_config.settings.download_url_zip_codes_org

    try:
        file_resp = requests.get(
            url=url,
            allow_redirects=True,
            stream=True,
            timeout=io_config.settings.download_timeout,
        )

        if file_resp.status_code != 200:
            # ERROR.00.906 Unexpected response status code='{status_code}'
            io_utils.terminate_fatal(
                glob.ERROR_00_906.replace("{status_code}", str(file_resp.status_code))
            )

        # INFO.00.058 The connection to the Zip Code Database file '{filename}'
        # on the Zip Codes.org download page was successfully established
        io_utils.progress_msg(
            glob.INFO_00_058.replace(
                "{filename}", io_config.settings.download_file_zip_codes_org_xlsx
            )
        )

        if not os.path.isdir(io_config.settings.download_work_dir):
            os.makedirs(io_config.settings.download_work_dir)

        filename_xls = os.path.join(
            io_config.settings.download_work_dir.replace("/", os.sep),
            io_config.settings.download_file_zip_codes_org_xlsx,
        )

        no_chunks = 0

        with open(filename_xls, "wb") as file_zip:
            for chunk in file_resp.iter_content(
                chunk_size=io_config.settings.download_chunk_size
            ):
                file_zip.write(chunk)
                no_chunks += 1

        # INFO.00.023 From the file '{filename}' {no_chunks} chunks were downloaded
        io_utils.progress_msg(
            glob.INFO_00_023.replace(
                "{filename}", io_config.settings.download_file_zip_codes_org_xlsx
            ).replace("{no_chunks}", str(no_chunks))
        )

    except ConnectionError:
        # ERROR.00.905 Connection problem with url='{url}'
        io_utils.terminate_fatal(glob.ERROR_00_905.replace("{url}", url))
    except TimeoutError:
        # ERROR.00.909 Timeout after '{timeout}' seconds with url='{url}
        io_utils.terminate_fatal(
            glob.ERROR_00_909.replace(
                "{timeout}", str(io_config.settings.download_timeout)
            ).replace("{url}", url)
        )

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load airports.
# ------------------------------------------------------------------
def load_airport_data() -> None:
    """Load airports."""
    logger.debug(io_glob.LOGGER_START)

    _load_airport_data()

    _load_runway_data()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load aviation occurrence categories.
# ------------------------------------------------------------------
def load_aviation_occurrence_categories() -> None:
    """Load aviation occurrence categories."""
    logger.debug(io_glob.LOGGER_START)

    _load_aviation_occurrence_categories()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load country and state data.
# ------------------------------------------------------------------
def load_country_state_data() -> None:
    """Load country and state data."""
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------

    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    # ------------------------------------------------------------------
    # Load the country data.
    # ------------------------------------------------------------------

    # INFO.00.059 Load country data
    io_utils.progress_msg(glob.INFO_00_059)
    io_utils.progress_msg("-" * 80)
    _load_country_data(
        conn_pg,
        cur_pg,
    )

    # ------------------------------------------------------------------
    # Load the state data.
    # ------------------------------------------------------------------

    io_utils.progress_msg("-" * 80)

    # INFO.00.060 Load state data
    io_utils.progress_msg(glob.INFO_00_060)
    io_utils.progress_msg("-" * 80)
    _load_state_data(
        conn_pg,
        cur_pg,
    )

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------

    utils.upd_io_processed_files(
        io_config.settings.download_file_countries_states_json, cur_pg
    )

    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load sequence of events sequence data.
# ------------------------------------------------------------------
def load_sequence_of_events() -> None:
    """Load sequence of events sequence data."""
    logger.debug(io_glob.LOGGER_START)

    _load_sequence_of_events()

    # pylint: disable=R0801
    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load simplemaps data.
# ------------------------------------------------------------------
def load_simplemaps_data() -> None:
    """Load simplemaps data."""
    logger.debug(io_glob.LOGGER_START)

    # ------------------------------------------------------------------
    # Start processing.
    # ------------------------------------------------------------------
    conn_pg, cur_pg = db_utils.get_postgres_cursor()

    # ------------------------------------------------------------------
    # Delete existing data.
    # ------------------------------------------------------------------
    cur_pg.execute(
        """
    DELETE FROM io_lat_lng
    WHERE SOURCE IN (%s, %s);
    """,
        (
            glob.SOURCE_SM_US_CITIES,
            glob.SOURCE_SM_US_ZIP_CODES,
        ),
    )

    if cur_pg.rowcount > 0:
        conn_pg.commit()

        # INFO.00.063 Processed data source '{data_source}'
        io_utils.progress_msg(
            glob.INFO_00_063.replace("{data_source}", glob.SOURCE_SM_US_CITIES)
        )
        io_utils.progress_msg(
            glob.INFO_00_063.replace("{data_source}", glob.SOURCE_SM_US_ZIP_CODES)
        )
        io_utils.progress_msg(f"Number rows deleted  : {str(cur_pg.rowcount):>8}")
        io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load zip data from a US city file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_zips_from_us_cities(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load zip data from a US zip code file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_zips_from_us_zips(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Load city data from a US city file.
    # ------------------------------------------------------------------
    _load_simplemaps_data_cities_from_us_cities(conn_pg, cur_pg)

    io_utils.progress_msg("-" * 80)

    # ------------------------------------------------------------------
    # Delete and insert averaged data.
    # ------------------------------------------------------------------
    _load_table_io_lat_lng_average(conn_pg, cur_pg)

    # ------------------------------------------------------------------
    # Finalize processing.
    # ------------------------------------------------------------------
    cur_pg.close()
    conn_pg.close()

    logger.debug(io_glob.LOGGER_END)


# ------------------------------------------------------------------
# Load ZIP Code Database data.
# ------------------------------------------------------------------
def load_zip_codes_org_data() -> None:
    """Load ZIP Code Database data."""
    logger.debug(io_glob.LOGGER_START)

    _load_zip_codes_org_data()

    logger.debug(io_glob.LOGGER_END)
